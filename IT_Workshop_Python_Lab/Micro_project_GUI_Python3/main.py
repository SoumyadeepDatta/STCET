# -- coding: utf-8 --


from tkinter import *
from tkinter.ttk import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk,Image
import openpyxl
from openpyxl import Workbook

file=Workbook()
file['Sheet'].title="Admission Form"
sheet=file.active

sheet.column_dimensions['A'].width=15
sheet.column_dimensions['B'].width=15
sheet.column_dimensions['C'].width=7
sheet.column_dimensions['D'].width=25
sheet.column_dimensions['E'].width=25
sheet.column_dimensions['F'].width=12
sheet.column_dimensions['G'].width=15
sheet.column_dimensions['H'].width=17
sheet.column_dimensions['I'].width=30
sheet.column_dimensions['J'].width=50
sheet.column_dimensions['K'].width=50
sheet.column_dimensions['L'].width=14
sheet.column_dimensions['M'].width=14
sheet.column_dimensions['N'].width=14
sheet.column_dimensions['O'].width=14
sheet.column_dimensions['P'].width=14
sheet.column_dimensions['Q'].width=14
sheet.column_dimensions['R'].width=16
sheet.column_dimensions['S'].width=18
sheet.column_dimensions['T'].width=16

sheet.cell(row=1, column=1).value = "First Name"  
sheet.cell(row=1, column=2).value = "Last Name"    
sheet.cell(row=1, column=3).value = "Gender" 
sheet.cell(row=1, column=4).value = "Father's Name"  
sheet.cell(row=1, column=5).value = "Mother's Name" 
sheet.cell(row=1, column=6).value = "Date of Birth"   
sheet.cell(row=1, column=7).value = "Phone No."   
sheet.cell(row=1, column=8).value = "Father's Phone No."     
sheet.cell(row=1, column=9).value = "E-mail"  
sheet.cell(row=1, column=10).value = "Current Address"
sheet.cell(row=1, column=11).value = "Permanent Address"
sheet.cell(row=1, column=12).value = "Department" 
sheet.cell(row=1, column=13).value = "Class 10 Board"  
sheet.cell(row=1, column=14).value = "Class 10 Marks"
sheet.cell(row=1, column=15).value = "Class 12 Board"  
sheet.cell(row=1, column=16).value = "Class 12 Marks"
sheet.cell(row=1, column=17).value = "Physics Marks"  
sheet.cell(row=1, column=18).value = "Chemistry Marks" 
sheet.cell(row=1, column=19).value = "Mathematics Marks" 
sheet.cell(row=1, column=20).value = "PCM Percentage" 

file.save('AdmissionForm.xlsx') 

root=Tk()
root.geometry("1000x800")
root.config(bg='light blue')
root.resizable(width=False, height=False)
root.title("Welcome to St.Thomas' College of Engineering & Technology!")

logo=Image.open("Logo.jpg")
resized_logo=logo.resize((340,360),Image.ANTIALIAS)
new_logo=ImageTk.PhotoImage(resized_logo)
label_logo=Label(image=new_logo)
label_logo.place(x=230,y=150)

Label(root, text="ST. THOMAS' COLLEGE OF ENGINEERING AND TECHNOLOGY",font='Britanic 18 bold').place(x=70,y=30)
Label(root, text='ADDRESS - 4, Diamond Harbour Road, Kidderore, Kolkata-7000023',font='Arial 12 italic').place(x=155,y=538)
Label(root, text='PHONE - 2448-1081/1082',font='Arial 12 italic').place(x=300,y=580)
Label(root, text='EMAIL-ID - www.stcet.org',font='Arial 12 italic').place(x=310,y=620)

my_img=Image.open("UserPhoto.jpg")
resized=my_img.resize((120,180),Image.ANTIALIAS)
new_pic=ImageTk.PhotoImage(resized)
my_label=Label(image=new_pic)
my_label.place(x=810,y=25)     
photo=Label(root, text="Photo to be inserted here.", font="Arial 12").place(x=780,y=220)

Label(root, text='Your Signature.',font='Arial 12').place(x=810,y=410)
Entry(root,width=40).place(x=750,y=380)

            
def page1():
    win1=Toplevel(root)
    win1.title("STCET Admission Form (Step 1 of 2)")
    win1.geometry("1000x800")
    win1.config(bg='light blue')
    win1.resizable(width=False, height=False)
     
    def page2():
        win2 = Toplevel(win1)
        win2.title("STCET Admission Form (Step 2 of 2)")
        win2.geometry("1000x800")
        win2.config(bg='light blue')
        win2.resizable(width=False, height=False)

        class10_var=StringVar()
        class12_var=StringVar()
        Phy_var=StringVar()
        Chem_var=StringVar()
        Maths_var=StringVar()
        pcm_avg=IntVar()
        pcm_avg.set(0.0)

        Label(win2, text="Educational Details", font='Georgia 23', relief=GROOVE).grid(row=0, column=0, padx=50, pady=20)
        Label(win2, text="Department",font='Arial 12 bold').grid(row=3, column=0, padx=20, pady=20)
        Label(win2, text='Class 10',font='Arial 12 bold').grid(row=4, column=0, padx=20, pady=20)
        Label(win2, text='Percentage',font='Arial 12 bold').grid(row=4, column=2, padx=20, pady=20)
        Entry(win2, textvariable=class10_var).grid(row=4, column=3, padx=20, pady=20)
        Label(win2, text='%').grid(row=4, column=4, padx=20, pady=20)
        Label(win2, text='Class 12',font='Arial 12 bold').grid(row=5, column=0, padx=20, pady=20)
        Label(win2, text='Percentage',font='Arial 12 bold').grid(row=5, column=2, padx=20, pady=20)
        Entry(win2, textvariable=class12_var).grid(row=5, column=3, padx=20, pady=20)
        Label(win2, text='%').grid(row=5, column=4, padx=20, pady=20)
        Label(win2, text='Class 12 PCM Marks', font='Georgia 23', relief=GROOVE).grid(row=6, column=0, padx=50, pady=20)
        Label(win2, text='Physics',font='Arial 12 bold').grid(row=7, column=0, padx=20, pady=20)
        Label(win2, text='Chemistry',font='Arial 12 bold').grid(row=8, column=0, padx=20, pady=20)
        Label(win2, text='Mathematics',font='Arial 12 bold').grid(row=9, column=0, padx=20, pady=20)
        Label(win2, text='PCM percentage',font='Arial 12 bold').grid(row=10, column=0, padx=20, pady=20)
        Label(win2, text='%').grid(row=10, column=2, padx=20, pady=20)
        Entry(win2, textvariable=Phy_var).grid(row=7, column=1, padx=20, pady=20)
        Entry(win2, textvariable=Chem_var).grid(row=8, column=1, padx=20, pady=20)
        Entry(win2, textvariable=Maths_var).grid(row=9, column=1, padx=20, pady=20)
        Entry(win2, textvariable=pcm_avg).grid(row=10, column=1, padx=20, pady=20)

        options=['DEPT.', 'CSE', 'IT', 'ECE', 'EE']
        clicked=StringVar()
        clicked.set('DEPT')
        drop=OptionMenu(win2, clicked, *options)
        drop.grid(row=3, column=1, padx=20, pady=20)

        board10 = ['BOARD', 'ICSE', 'CBSE', 'WBSE']
        b10 = StringVar()
        b10.set('BOARD')
        drop1 = OptionMenu(win2, b10, *board10)
        drop1.grid(row=4, column=1, padx=20, pady=20)

        board12 = ['BOARD', 'ISC', 'CBSE', 'WBSE']
        b12 = StringVar()
        b12.set('BOARD')
        drop2 = OptionMenu(win2, b12, *board12)
        drop2.grid(row=5, column=1, padx=20, pady=20)
        
        def resetPage2():
            result=messagebox.askquestion("CLEAR","Are you sure you want to reset?")
            if(result=='yes'):
                class10_var.set("")
                class12_var.set("")
                Phy_var.set("")
                Chem_var.set("")
                Maths_var.set("")
                pcm_avg.set(0.0)
                clicked.set('DEPT')
                b10.set('BOARD')
                b12.set('BOARD')
     
        def average():
            p=Phy_var.get()
            c=Chem_var.get()
            m=Maths_var.get()
            t=int(p)+int(c)+int(m)
            a=t/3
            pcm_avg.set(a)
            
        def store():
            firstname=f_name.get()
            lastname=l_name.get()
            gender_c=x.get()
            fathername=father_name.get()
            mothername=mother_name.get()
            DOB1=day_data.get()
            DOB2=month_data.get()
            DOB3=year_data.get()
            DOB=DOB1+"/"+DOB2+"/"+DOB3
            phno=phone_no_c.get()
            phone_number="+91"+phno
            fphno=phone_no_father.get()
            father_phone_no="+91"+fphno
            email_id_entry=str(mail_id.get())
            eas=str(lmail.get())
            fmail=email_id_entry+eas
            curr_add=current_ad.get()
            per_add=permanent_ad.get()
            dept=clicked.get()
            board_10=b10.get()
            board_12=b12.get()
            class_10_per=class10_var.get()
            class_12_per=class12_var.get()
            Phy_marks=Phy_var.get()
            Chem_marks=Chem_var.get()
            Maths_marks=Maths_var.get()
            pcm_per=pcm_avg.get()
            if(firstname=="" or lastname=="" or gender_c==0 or fathername=="" or mothername=="" or DOB1=="" or DOB2=="" or DOB3=="" or  phno=="" or fphno=="" or email_id_entry=="" or eas=='Domain_name' or curr_add=="" or per_add=="" or dept=='DEPT' or board_10=='BOARD' or board_12=='BOARD' or class_10_per=="" or class_12_per=="" or Phy_marks=="" or Chem_marks=="" or Maths_marks=="" or pcm_per==0.0):
                print('Error!')
                messagebox.showerror("Error :(","Please fill up all the fields!")
                f_name.set("")
                l_name.set("")
                x.set(0)
                father_name.set("")
                mother_name.set("")
                phone_no_c.set("")
                phone_no_father.set("")
                mail_id.set("")
                lmail.set("Domain_name")
                current_ad.set("")
                permanent_ad.set("")
                class10_var.set("")
                class12_var.set("")
                Phy_var.set("")
                Chem_var.set("")
                Maths_var.set("")
                pcm_avg.set(0.0)
                clicked.set('DEPT')
                b10.set('BOARD')
                b12.set('BOARD')
            elif(len(phone_number)!=13 or len(father_phone_no)!=13):
                print('Error!')
                messagebox.showerror("Error :(","The entered phone number is incorrect!")
                f_name.set("")
                l_name.set("")
                x.set(0)
                father_name.set("")
                mother_name.set("")
                phone_no_c.set("")
                phone_no_father.set("")
                mail_id.set("")
                lmail.set("Domain_name")
                current_ad.set("")
                permanent_ad.set("")
                class10_var.set("")
                class12_var.set("")
                Phy_var.set("")
                Chem_var.set("")
                Maths_var.set("")
                pcm_avg.set(0.0)
                clicked.set('DEPT')
                b10.set('BOARD')
                b12.set('BOARD')
            elif(len(DOB1)!=2 or len(DOB2)!=2 or len(DOB3)!=4):
                print('Error!')
                messagebox.showerror("Error :(","The entered date of birth format is incorrect!")
                f_name.set("")
                l_name.set("")
                x.set(0)
                father_name.set("")
                mother_name.set("")
                phone_no_c.set("")
                phone_no_father.set("")
                mail_id.set("")
                lmail.set("Domain_name")
                current_ad.set("")
                permanent_ad.set("")
                class10_var.set("")
                class12_var.set("")
                Phy_var.set("")
                Chem_var.set("")
                Maths_var.set("")
                pcm_avg.set(0.0)
                clicked.set('DEPT')
                b10.set('BOARD')
                b12.set('BOARD')
            else:
                result=messagebox.askquestion("SUBMIT","Are you sure you want to enter the data?")
                if(result=='yes'):
                    print('Done')
                    file=openpyxl.load_workbook('AdmissionForm.xlsx')
                    sheet=file.active
                
                    current_row=sheet.max_row
                    current_column=sheet.max_column
                    
                    sheet.cell(row=current_row + 1, column=1).value=firstname
                    sheet.cell(row=current_row + 1, column=2).value=lastname
                    if (gender_c==1):
                        sheet.cell(row=current_row + 1, column=3).value="Male"
                    elif (gender_c==2):
                        sheet.cell(row=current_row + 1, column=3).value="Female"
                    else:
                        sheet.cell(row=current_row + 1, column=3).value="Other"
                    sheet.cell(row=current_row + 1, column=4).value=fathername
                    sheet.cell(row=current_row + 1, column=5).value=mothername
                    sheet.cell(row=current_row + 1, column=6).value=DOB
                    sheet.cell(row=current_row + 1, column=7).value=phone_number
                    sheet.cell(row=current_row + 1, column=8).value=father_phone_no
                    sheet.cell(row=current_row + 1, column=9).value=fmail
                    sheet.cell(row=current_row + 1, column=10).value=curr_add
                    sheet.cell(row=current_row + 1, column=11).value=per_add
                    sheet.cell(row=current_row + 1, column=12).value=dept
                    sheet.cell(row=current_row + 1, column=13).value=board_10
                    sheet.cell(row=current_row + 1, column=14).value=board_12
                    sheet.cell(row=current_row + 1, column=15).value=class_10_per
                    sheet.cell(row=current_row + 1, column=16).value=class_12_per
                    sheet.cell(row=current_row + 1, column=17).value=Phy_marks
                    sheet.cell(row=current_row + 1, column=18).value=Chem_marks
                    sheet.cell(row=current_row + 1, column=19).value=Maths_marks
                    sheet.cell(row=current_row + 1, column=20).value=pcm_per
                
                    file.save('AdmissionForm.xlsx')
                    
        btnTotal=Button(win2, text="CALCULATE PERCENTAGE", command=average)
        btnTotal.grid(row=9, column=2, padx=20, pady=20)
    
        btn1=Button(win2, text="STORE DATA", command=store)
        btn1.grid(row=7, column=3, padx=20, pady=20)

        btn2=Button(win2, text="RESET", command=resetPage2)
        btn2.grid(row=9, column=3, padx=20, pady=20)

        btn3=Button(win2, text="CANCEL", command=win2.destroy)
        btn3.grid(row=11, column=3, padx=20, pady=20)
        
        var1=IntVar()
        var1.set(0)
        Checkbutton(win2,text='I hereby declare that all the information I have entered is true to the best of my knowlege and belief. I am liable for the discrepancies in data, if any.',variable=var1).place(x=60,y=600)
    
    def resetPage1():
        clear=messagebox.askquestion('CLEAR','Are you sure you want to reset?')
        if(clear=='yes'):
            f_name.set("")
            l_name.set("")
            x.set(0)
            father_name.set("")
            mother_name.set("")
            phone_no_c.set("")
            phone_no_father.set("")
            mail_id.set("")
            lmail.set("Domain_name")
            current_ad.set("")
            permanent_ad.set("")
        
    f_name=StringVar()
    l_name=StringVar()
    father_name=StringVar()
    day_data=StringVar()
    month_data=StringVar()
    year_data=StringVar()
    mother_name=StringVar()
    phone_no_c=StringVar()
    phone_no_father=StringVar()
    current_ad=StringVar()
    permanent_ad=StringVar()
    mail_id=StringVar()
    
    head1=Label(win1, text="STCET",font='Georgia 60 bold',relief=RAISED).grid(row=1, column=2, padx=20, pady=20)
    head2=Label(win1, text="Admission Form",font='Granolia 23 bold',relief=GROOVE).grid(row=2, column=2, padx=20, pady=20)
    
    fname=Label(win1, text="First Name: ",font="Arial 12 bold").grid(row=3, column=0, padx=20, pady=20)
    lname=Label(win1, text="Last Name: ",font="Arial 12 bold").grid(row=3, column=2, padx=20, pady=20)
    fname_field=Entry(win1, textvariable=f_name, width=30).grid(row=3, column=1, padx=20, pady=20)
    lname_field=Entry(win1, textvariable=l_name, width=30).grid(row=3, column=3, padx=20, pady=20)
    gender=Label(win1, text="Gender: ",font="Arial 12 bold").grid(row=4, column=0, padx=20, pady=20)
    x = IntVar()
    x.set(0)
    rad1=Radiobutton(win1, text="Male", variable=x, value=1).grid(row=4, column=1, padx=5, pady=5)
    rad2=Radiobutton(win1, text="Female", variable=x, value=2).grid(row=4, column=2, padx=5, pady=5)
    rad3=Radiobutton(win1, text="Other", variable=x, value=3).grid(row=4, column=3, padx=5, pady=5)
    father=Label(win1, text="Father's Name: ",font="Arial 12 bold").grid(row=5,column=0, padx=20, pady=20)
    father_field=Entry(win1, textvariable=father_name,width=30).grid(row=5, column=1, padx=20, pady=20)
    dob=Label(win1, text="Date Of Birth (DD/MM/YYYY) :",font="Arial 12 bold").grid(row=5, column=2, padx=20, pady=20)
    d_field=Entry(win1, textvariable=day_data,width=10).place(x=720,y=360)
    s1=Label(win1,text="/").place(x=790,y=360)
    m_field=Entry(win1, textvariable=month_data,width=10).place(x=803,y=360)
    s2=Label(win1,text="/").place(x=873,y=360)
    y_field=Entry(win1, textvariable=year_data,width=16).place(x=885,y=360)
    mother=Label(win1, text="Mother's Name: ",font="Arial 12 bold").grid(row=6, column=0, padx=20, pady=20)
    mother_field=Entry(win1,textvariable=mother_name, width=30).grid(row=6, column=1, padx=20, pady=20)
    contact=Label(win1, text="Contact No.: ",font="Arial 12 bold").grid(row=7, column =0, padx=20, pady=20)
    contact_field=Entry(win1, textvariable=phone_no_c,width=30).grid(row=7, column=1, padx=20, pady=20)
    ct=Label(win1, text="Father's Contact no.:",font="Arial 12 bold").grid(row=8, column=0 , padx=20, pady=20)
    ct_field=Entry(win1, textvariable=phone_no_father,width=30).grid(row=8,column=1,padx=20,pady=20)
    email=Label(win1, text="Email id: ",font="Arial 12 bold").grid(row=9, column =0, padx=20, pady=20)
    email_field=Entry(win1, textvariable=mail_id, width=30).grid(row=9, column=1, padx=20, pady=20)
    ops=['Domain_name','@gmail.com','@yahoo.com','@hotmail.com','@rediffmail.com']
    lmail=StringVar()
    lmail.set('Domain_name')
    e_drop=OptionMenu(win1, lmail, *ops)
    e_drop.grid(row=9,column=2)
    cadd = Label(win1, text="Current Address: ",font="Arial 12 bold").grid(row=6, column=2, padx=20, pady=20)
    cadd_field = Entry(win1, textvariable=current_ad, width=30).grid(row=6, column=3, padx=20, pady=20)
    padd = Label(win1, text="Permanent Address: ",font="Arial 12 bold").grid(row=7, column=2, padx=20, pady=20)
    padd_field = Entry(win1, textvariable=permanent_ad, width=30).grid(row=7, column=3, padx=20, pady=20)

    def cont2():
        if(messagebox.askokcancel("Confirmation", "Do you want to continue?")):
            page2()
        else:
            pass

    def close():
        if (messagebox.askokcancel("Confirmation", "Want to cancel?")):
            root.destroy()
        else:
            pass

    Button(win1, text="RESET", command=resetPage1).grid(row=10, column=1, padx=20, pady=20)
    Button(win1, text="CANCEL", command=close).grid(row=10, column=2, padx=20, pady=20)
    Button(win1, text="CONTINUE", command=cont2).grid(row=10, column=3, padx=20, pady=20)
    
def cont1():
    if(messagebox.askokcancel("Confirmation", "Do you want to continue?")):
        page1()
    else:
        pass
Button(root, text="CONTINUE", command=cont1).grid(row=10, column=3, padx=800, pady=600)
mainloop()
